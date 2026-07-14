"""Excel tracker I/O."""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import openpyxl

from .notice import FIELD_NAMES, HEADERS, AuctionRow


def _sort_key(row: AuctionRow) -> tuple[int, str]:
    if row.tender_no is not None:
        try:
            return (int(row.tender_no), row.isin)
        except (TypeError, ValueError):
            pass
    return (0, row.isin)


def _write_row(ws: Any, r_idx: int, row: AuctionRow) -> None:
    for c_idx, name in enumerate(FIELD_NAMES, 1):
        ws.cell(row=r_idx, column=c_idx, value=getattr(row, name))


def _read_existing_keys(ws: Any) -> set[tuple[str | None, str | None]]:
    tender_col = FIELD_NAMES.index("tender_no") + 1
    isin_col = FIELD_NAMES.index("isin") + 1
    return {
        (ws.cell(row=r, column=tender_col).value, ws.cell(row=r, column=isin_col).value)
        for r in range(2, ws.max_row + 1)
    }


def build(out_path: Path, rows: list[AuctionRow]) -> int:
    rows = sorted(rows, key=_sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auction Results"
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    for i, row in enumerate(rows, start=2):
        _write_row(ws, i, row)

    wb.save(out_path)
    print(f"wrote {len(rows)} rows to {out_path}")
    return len(rows)


def append(out_path: Path, rows: list[AuctionRow]) -> int:
    wb = openpyxl.load_workbook(out_path)
    if "Auction Results" not in wb.sheetnames:
        print(f"error: '{out_path}' has no 'Auction Results' sheet", file=sys.stderr)
        return 0
    ws = wb["Auction Results"]

    existing_keys = _read_existing_keys(ws)
    new_rows = sorted((r for r in rows if r.key not in existing_keys), key=_sort_key)
    skipped = len(rows) - len(new_rows)

    next_row = ws.max_row + 1
    for row in new_rows:
        _write_row(ws, next_row, row)
        next_row += 1

    wb.save(out_path)
    print(f"added {len(new_rows)} rows to {out_path} ({skipped} already present)")
    return len(new_rows)
