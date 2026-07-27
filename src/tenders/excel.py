"""CLI orchestration for parsing PDFs into Excel."""

from __future__ import annotations

import argparse
import sys
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import openpyxl
from rich.console import Console
from rich.progress import (
    BarColumn,
    Progress,
    TaskID,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
)

from .notice import FIELD_NAMES, HEADERS, AuctionRow, ParseError, parse_pdf

console = Console()


def discover_pdfs(paths: Iterable[str], recursive: bool = False) -> list[Path]:
    found: set[Path] = set()
    for raw in paths:
        p = Path(raw)
        if p.is_dir():
            pattern = "**/*.pdf" if recursive else "*.pdf"
            matches = sorted(p.glob(pattern))
            if not matches:
                print(f"warning: {p}: no PDFs found", file=sys.stderr)
            found.update(matches)
        elif p.is_file():
            found.add(p)
        else:
            print(f"warning: {p}: path does not exist, skipping", file=sys.stderr)
    return sorted(found)


def collect_rows(
    pdf_paths: list[Path],
    progress: Progress | None = None,
    task_id: TaskID | None = None,
) -> list[AuctionRow]:
    rows: list[AuctionRow] = []
    for path in pdf_paths:
        try:
            rows.extend(parse_pdf(path))
        except ParseError as exc:
            print(f"warning: {path.name}: skipped — {exc}", file=sys.stderr)
        if progress is not None and task_id is not None:
            progress.update(task_id, advance=1)
    return rows


def _sort_key(row: AuctionRow) -> tuple[int, str]:
    if row.tender_no is not None:
        try:
            return (int(row.tender_no), row.isin)
        except (TypeError, ValueError):
            pass
    return (0, row.isin)


def _read_existing_keys(ws: Any) -> set[tuple[str | None, str | None]]:
    tender_col = FIELD_NAMES.index("tender_no") + 1
    isin_col = FIELD_NAMES.index("isin") + 1
    return {
        (ws.cell(row=r, column=tender_col).value, ws.cell(row=r, column=isin_col).value)
        for r in range(2, ws.max_row + 1)
    }


def _build(out_path: Path, rows: list[AuctionRow]) -> int:
    rows = sorted(rows, key=_sort_key)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auction Results"
    ws.append(HEADERS)
    for row in rows:
        ws.append([getattr(row, name) for name in FIELD_NAMES])
    wb.save(out_path)
    print(f"wrote {len(rows)} rows to {out_path}")
    return len(rows)


def _append(out_path: Path, rows: list[AuctionRow]) -> int:
    wb = openpyxl.load_workbook(out_path)
    if "Auction Results" not in wb.sheetnames:
        print(f"error: '{out_path}' has no 'Auction Results' sheet", file=sys.stderr)
        return 0
    ws = wb["Auction Results"]

    existing_keys = _read_existing_keys(ws)
    new_rows = sorted((r for r in rows if r.key not in existing_keys), key=_sort_key)
    skipped = len(rows) - len(new_rows)

    for row in new_rows:
        ws.append([getattr(row, name) for name in FIELD_NAMES])

    wb.save(out_path)
    print(f"added {len(new_rows)} rows to {out_path} ({skipped} already present)")
    return len(new_rows)


def main(args: argparse.Namespace) -> None:
    pdf_paths = discover_pdfs(args.paths, recursive=args.recursive)
    if not pdf_paths:
        print("error: no PDF files found in the given paths", file=sys.stderr)
        raise SystemExit(1)

    with Progress(
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("  Parsing PDFs", total=len(pdf_paths))
        rows = collect_rows(pdf_paths, progress=progress, task_id=task)

    if args.new or not args.tracker.exists():
        n = _build(args.tracker, rows)
        if not n:
            raise SystemExit(1)
    else:
        _append(args.tracker, rows)
